import os
import torch
from torch import nn
from transformers import AutoModelForCausalLM, AutoTokenizer, AutoConfig, default_data_collator, get_linear_schedule_with_warmup, Trainer, get_scheduler, DataCollatorWithPadding
from transformers.modeling_outputs import TokenClassifierOutput
from huggingface_hub import snapshot_download
import math
import pandas as pd
import json
import pickle
import numpy as np
from utils import ROOT_DIR
from datasets import load_dataset
from collections import Counter
from sklearn.model_selection import train_test_split
import random
from openpyxl import Workbook, load_workbook
from torch.utils.data import DataLoader
from tqdm import tqdm
from datasets import Dataset
from torch.optim import AdamW


def read_excel_file_responses(file_path):
    column1_strings = []  # List to store strings from the first column
    column2_strings = []  # List to store strings from the second column
    try:
        # Load the workbook
        workbook = load_workbook(file_path)
        # Select the first sheet
        sheet = workbook.active
        # Iterate over rows in the first column and append strings to the list
        for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row, values_only=True):
            column1_strings.append(row[0])
        # Iterate over rows in the second column and append strings to the list
        for row in sheet.iter_rows(min_row=1, min_col=2, max_col=2, max_row=sheet.max_row, values_only=True):
            column2_strings.append(row[0])
    except Exception as e:
        print("An error occurred:", e)
    return column1_strings, column2_strings


def load_rewardbench_chat_OOD_reasoning():

    def generate_random_data(samples, labels, num_samples):
        # Create a permutation
        permutation = list(range(len(samples)))
        random.shuffle(permutation)
        
        # Select samples and labels based on the permutation and num_samples
        random_samples = [samples[i] for i in permutation[:num_samples]]
        random_labels = [labels[i] for i in permutation[:num_samples]]
        return random_samples, random_labels

    def balance_binary_classes(samples, labels):
        class_0_samples = [sample for sample, label in zip(samples, labels) if label == 0]
        class_1_samples = [sample for sample, label in zip(samples, labels) if label == 1]
        
        min_samples = min(len(class_0_samples), len(class_1_samples))
        
        balanced_samples = class_0_samples[:min_samples] + class_1_samples[:min_samples]
        balanced_labels = [0] * min_samples + [1] * min_samples
        
        return balanced_samples, balanced_labels

    # Load the training dataset
    dataset = load_dataset("allenai/reward-bench")

    train_prompt = dataset['filtered']['prompt']
    train_chosen = dataset['filtered']['chosen']
    train_rejected = dataset['filtered']['rejected']
    train_subset = dataset['filtered']['subset']

    train_filtered_prompt = [train_prompt[i] for i in range(len(train_subset)) if train_subset[i] in ['alpacaeval-easy', 'alpacaeval-length', 'alpacaeval-hard', 'mt-bench-easy', 'mt-bench-med']]
    train_filtered_chosen = [train_chosen[i] for i in range(len(train_subset)) if train_subset[i] in ['alpacaeval-easy', 'alpacaeval-length', 'alpacaeval-hard', 'mt-bench-easy', 'mt-bench-med']]
    train_filtered_rejected = [train_rejected[i] for i in range(len(train_subset)) if train_subset[i] in ['alpacaeval-easy', 'alpacaeval-length', 'alpacaeval-hard', 'mt-bench-easy', 'mt-bench-med']]

    train_chosen = [
        f"<s>[INST] {train_filtered_prompt[i]} </s> [/INST] {train_filtered_chosen[i]}</s>"
        for i in range(len(train_filtered_prompt))
    ]
    train_chosen_labels = [0] * len(train_chosen)

    train_rejected = [
        f"<s>[INST] {train_filtered_prompt[i]} </s> [/INST] {train_filtered_rejected[i]}</s>"
        for i in range(len(train_filtered_prompt))
    ]
    train_rejected_labels = [1] * len(train_rejected)
    train_sentences = train_chosen + train_rejected
    train_labels = train_chosen_labels + train_rejected_labels

    val_indices = random.sample(range(len(train_sentences)), 50)
    val_sentences = [train_sentences[idx] for idx in val_indices]
    val_labels = [0] * len(val_sentences)
    train_sentences = [sample for i, sample in enumerate(train_sentences) if i not in val_indices]

    test_indices = random.sample(range(len(train_sentences)), 250)
    test_sentences = [train_sentences[idx] for idx in test_indices]
    test_labels = [0] * len(test_sentences)
    train_sentences = [sample for i, sample in enumerate(train_sentences) if i not in test_indices]
    train_labels = [0] * len(train_sentences)

    print('\n\ntrain_labels_IID', len(train_labels))
    print('test_labels_IID', len(test_labels))
    print('val_labels_IID', len(val_labels))
    print('test_labels_IID', test_labels[0:10])
    print('Unique IID train labels', {num: train_labels.count(num) for num in train_labels})
    print('Unique IID test labels', {num: test_labels.count(num) for num in test_labels})
    print('Unique IID val labels', {num: val_labels.count(num) for num in val_labels})
    
    ###Load generated-OOD dataset
    train_prompt = dataset['filtered']['prompt']
    train_chosen = dataset['filtered']['chosen']
    train_rejected = dataset['filtered']['rejected']
    train_subset = dataset['filtered']['subset']

    OOD_train_filtered_prompt = [train_prompt[i] for i in range(len(train_subset)) if train_subset[i] in ['hep-python', 'math-prm', 'hep-cpp', 'hep-java', 'hep-rust']]
    OOD_train_filtered_chosen = [train_chosen[i] for i in range(len(train_subset)) if train_subset[i] in ['hep-python', 'math-prm', 'hep-cpp', 'hep-java', 'hep-rust']]
    OOD_train_filtered_rejected = [train_rejected[i] for i in range(len(train_subset)) if train_subset[i] in ['hep-python', 'math-prm', 'hep-cpp', 'hep-java', 'hep-rust']]

    OOD_train_chosen = [
        f"<s>[INST] {OOD_train_filtered_prompt[i]} </s> [/INST] {OOD_train_filtered_chosen[i]}</s>"
        for i in range(len(OOD_train_filtered_prompt))
    ]
    OOD_train_chosen_labels = [0] * len(OOD_train_chosen)

    OOD_train_rejected = [
        f"<s>[INST] {OOD_train_filtered_prompt[i]} </s> [/INST] {OOD_train_filtered_rejected[i]}</s>"
        for i in range(len(OOD_train_filtered_prompt))
    ]
    OOD_train_sentences = OOD_train_chosen + OOD_train_rejected

    OOD_test_indices = random.sample(range(len(OOD_train_sentences)), 250)
    OOD_test_sentences = [OOD_train_sentences[idx] for idx in OOD_test_indices]
    OOD_test_labels = [1] * len(OOD_test_sentences)

    
    OOD_train_filtered_prompt1, OOD_train_filtered_chosen1 = read_excel_file_responses("data/data_ethical/OOD_samples_llama-3-70b-instruct_math_augmented_beaver-withresponses.xlsx")    
    OOD_train_filtered_prompt1 = OOD_train_filtered_prompt1[1:500]
    OOD_train_filtered_chosen1 = OOD_train_filtered_chosen1[1:500]

    OOD_train_filtered_prompt2, OOD_train_filtered_chosen2 = read_excel_file_responses("data/data_ethical/OOD_samples_llama-3-70b-instruct_code_augmented_beaver-withresponses.xlsx")    
    OOD_train_filtered_prompt2 = OOD_train_filtered_prompt2[1:500]
    OOD_train_filtered_chosen2 = OOD_train_filtered_chosen2[1:500]

    OOD_train_filtered_prompt3, OOD_train_filtered_chosen3 = read_excel_file_responses("data/data_ethical/OOD_samples_llama-3-70b-instruct_java_augmented_beaver-withresponses.xlsx")    
    OOD_train_filtered_prompt3 = OOD_train_filtered_prompt3[1:500]
    OOD_train_filtered_chosen3 = OOD_train_filtered_chosen3[1:500]

    OOD_train_filtered_prompt4, OOD_train_filtered_chosen4 = read_excel_file_responses("data/data_ethical/OOD_samples_llama-3-70b-instruct_rust_augmented_beaver-withresponses.xlsx")    
    OOD_train_filtered_prompt4 = OOD_train_filtered_prompt4[1:500]
    OOD_train_filtered_chosen4 = OOD_train_filtered_chosen4[1:500]

    OOD_train_filtered_prompt5, OOD_train_filtered_chosen5 = read_excel_file_responses("data/data_ethical/OOD_samples_llama-3-70b-instruct_cpp_augmented_beaver-withresponses.xlsx")    
    OOD_train_filtered_prompt5 = OOD_train_filtered_prompt5[1:500]
    OOD_train_filtered_chosen5 = OOD_train_filtered_chosen5[1:500]

    OOD_train_filtered_prompt = OOD_train_filtered_prompt1 + OOD_train_filtered_prompt2 + OOD_train_filtered_prompt3 + OOD_train_filtered_prompt4 + OOD_train_filtered_prompt5
    OOD_train_filtered_chosen = OOD_train_filtered_chosen1 + OOD_train_filtered_chosen2 + OOD_train_filtered_chosen3 + OOD_train_filtered_chosen4 + OOD_train_filtered_chosen5

    OOD_train_chosen = [
        f"<s>[INST] {OOD_train_filtered_prompt[i]} </s> [/INST] {OOD_train_filtered_chosen[i]}</s>"
        for i in range(len(OOD_train_filtered_prompt))
    ]
    OOD_train_chosen_labels = [0] * len(OOD_train_chosen)

    OOD_train_sentences = OOD_train_chosen
    OOD_train_labels = OOD_train_chosen_labels 

    OOD_val_indices = random.sample(range(len(OOD_train_sentences)), 200)
    OOD_val_sentences = [OOD_train_sentences[idx] for idx in OOD_val_indices]
    OOD_val_labels = [1] * len(OOD_val_sentences)
    OOD_train_sentences = [sample for i, sample in enumerate(OOD_train_sentences) if i not in OOD_val_indices]
    OOD_train_labels = [1] * len(OOD_train_sentences)
    OOD_train_sentences, OOD_train_labels = generate_random_data(OOD_train_sentences, OOD_train_labels, 1500)

    print('\n\nOOD_train_sentences', len(OOD_train_sentences))
    print('OOD_train_labels', len(OOD_train_labels))
    print('OOD_test_sentences', len(OOD_test_sentences))
    print('OOD_test_labels', len(OOD_test_labels))
    print('OOD_val_sentences', len(OOD_val_sentences))
    print('OOD_val_labels', len(OOD_val_labels))

    ###Merge IID+generated-OOD dataset
    train_sentences.extend(OOD_train_sentences)
    train_labels.extend(OOD_train_labels)
    test_sentences.extend(OOD_test_sentences)
    test_labels.extend(OOD_test_labels)
    # test_sentences = OOD_test_sentences
    # test_labels= OOD_test_labels
    val_sentences.extend(OOD_val_sentences)
    val_labels.extend(OOD_val_labels)
    
    print('\n\ntrain_labels_IID+OOD', len(train_labels))
    print('Unique train labels', {num: train_labels.count(num) for num in train_labels})
    print('Unique test labels', {num: test_labels.count(num) for num in test_labels})
    print('Unique val labels', {num: val_labels.count(num) for num in val_labels})
    print('test_labels', len(test_labels))
    print('val_labels', len(val_labels))

    return train_sentences, train_labels, test_sentences, test_labels, val_sentences, val_labels


class GPTRewardModel(nn.Module):
    def __init__(self, model_path):
        super().__init__()
        # model = AutoModelForCausalLM.from_pretrained(model_path
        config = AutoConfig.from_pretrained(model_path)
        config.pretraining_tp = 1
        model = AutoModelForCausalLM.from_pretrained(model_path, device_map='sequential', config=config,
                                                             torch_dtype=torch.float16,
                                                             low_cpu_mem_usage=True)
        self.config = model.config
        self.config.n_embd = self.config.hidden_size if hasattr(self.config, "hidden_size") else self.config.n_embd
        self.model = model
        self.transformer = model.model
        self.v_head = nn.Linear(self.config.n_embd, 1, bias=False)

        self.num_labels = 2 
        self.dropout = nn.Dropout(0.1) 
        self.classifier = nn.Linear(4096, self.num_labels)

        self.tokenizer = AutoTokenizer.from_pretrained(model_path)
        self.tokenizer.pad_token = self.tokenizer.unk_token
        self.PAD_ID = self.tokenizer(self.tokenizer.pad_token)["input_ids"][0]

    def get_device(self):
        return self.model.device

    def forward(
        self,
        input_ids=None,
        past_key_values=None,
        attention_mask=None,
        labels=None,
        position_ids=None,
    ):
        """
        input_ids, attention_mask: torch.Size([bs, seq_len])
        return: scores: List[bs]
        """
        bs = input_ids.shape[0]
        transformer_outputs = self.transformer(
            input_ids,
            past_key_values=past_key_values,
            attention_mask=attention_mask,
            position_ids=position_ids,
        )
        hidden_states = transformer_outputs[0]

        #Add custom layers
        sequence_output = self.dropout(hidden_states) #outputs[0]=last hidden state

        logits = self.classifier(sequence_output[:, -1, :]) # calculate losses

        loss = None
        if labels is not None:
            loss_fct = nn.CrossEntropyLoss()
            loss = loss_fct(logits.view(-1, self.num_labels), labels.view(-1))

        scores = []
        rewards = self.v_head(hidden_states).squeeze(-1)
        for i in range(bs):
            c_inds = (input_ids[i] == self.PAD_ID).nonzero()
            c_ind = c_inds[0].item() if len(c_inds) > 0 else input_ids.shape[1]
            scores.append(rewards[i, c_ind - 1])
        return TokenClassifierOutput(loss=loss, logits=logits, hidden_states=transformer_outputs.hidden_states,attentions=transformer_outputs.attentions)

## Load the model and tokenizer
reward_model = GPTRewardModel('/home/abbasm2/hugginface/hub/Llama-2-7b-chat-hf')
reward_tokenizer = reward_model.tokenizer
reward_tokenizer.truncation_side = "left"

# Note: The `token` parameter refers to your Hugging Face access token, required to download private or restricted models.
directory = snapshot_download(repo_id="berkeley-nest/Starling-RM-7B-alpha", token='enter_your_HF_token_here', local_dir='/home/abbasm2/hugginface/hub/Starling-RM-7B-alpha')
for fpath in os.listdir(directory):
    if fpath.endswith(".pt") or fpath.endswith("model.bin"):
        checkpoint = os.path.join(directory, fpath)
        break

reward_device = "cuda"
reward_batch_size = 16
max_length = 128
lr = 0.00015
num_epochs = 10

reward_model.load_state_dict(torch.load(checkpoint), strict=False)
reward_model = reward_model.to(reward_device)
reward_model.train()


for name, param in reward_model.named_parameters():
    if (name == 'classifier.weight' or name == 'classifier.bias'):
        param.requires_grad = True
    else:
        param.requires_grad = False
    print(name, param.requires_grad)
     

## Define the reward function
def get_reward(samples):
    """samples: List[str]"""
    input_ids = []
    attention_masks = []
    encodings_dict = reward_tokenizer(
        samples,
        truncation=True,
        max_length=2048,
        padding="max_length",
        return_tensors="pt",
    ).to(reward_device)
    input_ids = encodings_dict["input_ids"]
    attention_masks = encodings_dict["attention_mask"]
    mbs = reward_batch_size
    out = []
    out_class = []
    for i in range(math.ceil(len(samples) / mbs)):
        rewards, classes = reward_model(input_ids=input_ids[i * mbs : (i + 1) * mbs], attention_mask=attention_masks[i * mbs : (i + 1) * mbs])
        out.extend(rewards)
        out_class.extend(classes)
    return torch.hstack(out), torch.hstack(out_class)


all_train_sentences, all_train_labels, all_test_sentences, all_test_labels, all_val_sentences, all_val_labels = load_rewardbench_chat_OOD_reasoning()

dataset = {}

dataset["train"] = {'labels': all_train_labels[:len(all_train_labels)], 'text': all_train_sentences[:len(all_train_labels)]}
dataset["test"] = {'labels': all_test_labels[:len(all_test_labels)], 'text': all_test_sentences[:len(all_test_labels)]}
dataset["val"] = {'labels': all_val_labels[:len(all_val_labels)], 'text': all_val_sentences[:len(all_val_labels)]}

dataset["train"] = Dataset.from_dict(dataset["train"])
dataset["test"] = Dataset.from_dict(dataset["test"])
dataset["val"] = Dataset.from_dict(dataset["val"])

print('\n\n train: ', dataset["train"][0:3])
print('\n test: ', dataset["test"][0:3])
print('\n val: ', dataset["val"][0:3])

#tokenize
def tokenize_function(examples):
    return reward_tokenizer(examples["text"], max_length=128, truncation=True)

train_tokenized_datasets = dataset["train"].map(tokenize_function, batched=True)
test_tokenized_datasets = dataset["test"].map(tokenize_function, batched=True)
val_tokenized_datasets = dataset["val"].map(tokenize_function, batched=True)

# manually postprocess tokenized_dataset to prepare it for training.
train_tokenized_datasets.set_format("torch", columns=["input_ids", "attention_mask", "labels"])
test_tokenized_datasets.set_format("torch", columns=["input_ids", "attention_mask", "labels"])
val_tokenized_datasets.set_format("torch", columns=["input_ids", "attention_mask", "labels"])

data_collator = DataCollatorWithPadding(tokenizer=reward_tokenizer)

# create a smaller subset of the dataset as previously shown to speed up the fine-tuning:
train_dataset = train_tokenized_datasets.shuffle(seed=42).select(range(1916))
test_dataset = test_tokenized_datasets.shuffle(seed=42).select(range(500))
eval_dataset = val_tokenized_datasets.shuffle(seed=42).select(range(250))

train_dataloader = DataLoader(train_dataset, shuffle=True, batch_size=reward_batch_size, collate_fn=data_collator)
test_dataloader = DataLoader(test_dataset, batch_size=reward_batch_size, collate_fn=data_collator)
eval_dataloader = DataLoader(eval_dataset, batch_size=reward_batch_size, collate_fn=data_collator)

optimizer = AdamW(reward_model.parameters(), lr=4e-5)
num_training_steps = num_epochs * len(train_dataloader)
lr_scheduler = get_scheduler(
    "linear",
    optimizer=optimizer,
    num_warmup_steps=0,
    num_training_steps=num_training_steps,
)

progress_bar_train = tqdm(range(num_training_steps))
progress_bar_eval = tqdm(range(num_epochs * len(eval_dataloader)))


from datasets import load_metric
metric = load_metric("accuracy", trust_remote_code=True)
for epoch in range(num_epochs):
    reward_model.train()
    for batch in train_dataloader:
        batch = {k: v.to(reward_device) for k, v in batch.items()}
        outputs = reward_model(**batch)
        loss = outputs.loss
        loss.backward()

        optimizer.step()
        lr_scheduler.step()
        optimizer.zero_grad()
        progress_bar_train.update(1)

    reward_model.eval()
    for batch in eval_dataloader:
        batch = {k: v.to(reward_device) for k, v in batch.items()}
        with torch.no_grad():
            outputs = reward_model(**batch)

        logits = outputs.logits
        predictions = torch.argmax(logits, dim=-1)
        metric.add_batch(predictions=predictions, references=batch["labels"])
        progress_bar_eval.update(1)
        
    print('val_acc={}'.format(metric.compute()))


reward_model.eval()
for batch in test_dataloader:
    batch = {k: v.to(reward_device) for k, v in batch.items()}
    with torch.no_grad():
        outputs = reward_model(**batch)

    logits = outputs.logits
    predictions = torch.argmax(logits, dim=-1)
    metric.add_batch(predictions=predictions, references=batch["labels"])

print('\n\n final_test_acc={}'.format(metric.compute()))


