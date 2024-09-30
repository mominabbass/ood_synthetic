import pandas as pd
import json
import pickle
import numpy as np
from utils import ROOT_DIR
from datasets import load_dataset
from collections import Counter
from sklearn.model_selection import train_test_split
import random
from openpyxl import Workbook, load_workbook

def load_sst2():
    def process_raw_data_sst(lines):
        """from lines in dataset to two lists of sentences and labels respectively"""
        labels = []
        sentences = []
        for line in lines:
            labels.append(int(line[0]))
            sentences.append(line[2:].strip())
        return sentences, labels

    with open(f"{ROOT_DIR}/data/sst2/stsa.binary.train", "r") as f:
        train_lines = f.readlines()
    with open(f"{ROOT_DIR}/data/sst2/stsa.binary.test", "r") as f:
        test_lines = f.readlines()
    with open(f"{ROOT_DIR}/data/sst2/stsa.binary.dev", "r") as f:
        val_lines = f.readlines()
    train_sentences, train_labels = process_raw_data_sst(train_lines)
    test_sentences, test_labels = process_raw_data_sst(test_lines)
    val_sentences, val_labels = process_raw_data_sst(val_lines)
    return train_sentences, train_labels, test_sentences, test_labels, val_sentences, val_labels


def load_beavertails_unethical_OOD_gsm8k():
    def generate_random_data(samples, labels, num_samples):
        random.seed(42)

        # Create a permutation
        permutation = list(range(len(samples)))
        random.shuffle(permutation)
        
        # Select samples and labels based on the permutation and num_samples
        random_samples = [samples[i] for i in permutation[:num_samples]]
        random_labels = [labels[i] for i in permutation[:num_samples]]
        return random_samples, random_labels

    def balance_binary_classes(samples, labels):
        class_0_samples = [sample for sample, label in zip(samples, labels) if label == 0]
        class_1_samples = [sample for sample, label in zip(samples, labels) if label == 1]
        
        min_samples = min(len(class_0_samples), len(class_1_samples))
        
        balanced_samples = class_0_samples[:min_samples] + class_1_samples[:min_samples]
        balanced_labels = [0] * min_samples + [1] * min_samples
        
        return balanced_samples, balanced_labels

    # Load the training dataset
    train_data = load_dataset('PKU-Alignment/BeaverTails', split='330k_train')
    test_data = load_dataset('PKU-Alignment/BeaverTails', split='330k_test')

    train_prompts = train_data['prompt']
    train_response = train_data['response']
    train_categories = train_data['category']
    train_is_safe = train_data['is_safe']
    test_prompts = test_data['prompt']
    test_response = test_data['response']
    test_categories = test_data['category']
    test_is_safe = test_data['is_safe']

    # Extract prompts with "discrimination, stereotype, and injustice" category as True
    train_filtered_prompts = [train_prompts[i] for i in range(len(train_categories)) if train_categories[i]['non_violent_unethical_behavior']]
    train_filtered_responses = [train_response[i] for i in range(len(train_categories)) if train_categories[i]['non_violent_unethical_behavior']]
    train_filtered_prompt_final = []
    for i in range(len(train_filtered_prompts)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {train_filtered_prompts[i]}\nResponse: {train_filtered_responses[i]}"""
        train_filtered_prompt_final.append(formatted_entry)
    train_filtered_labels = [1] * len(train_filtered_prompt_final)
    train_safe_prompts = [train_prompts[i] for i in range(len(train_prompts)) if train_is_safe[i]][0:60000]
    train_safe_responses = [train_response[i] for i in range(len(train_response)) if train_is_safe[i]][0:60000]
    train_safe_prompt_final = []
    for i in range(len(train_safe_prompts)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {train_safe_prompts[i]}\nResponse: {train_safe_responses[i]}"""
        train_safe_prompt_final.append(formatted_entry)
    train_safe_labels = [0] * len(train_safe_prompt_final)
    train_sentences = train_filtered_prompt_final + train_safe_prompt_final
    train_labels = train_filtered_labels + train_safe_labels
    train_sentences, train_labels = balance_binary_classes(train_sentences, train_labels)
    train_sentences, train_labels = generate_random_data(train_sentences, train_labels, 17600)
    train_labels = [0] * len(train_sentences)

    val_indices = random.sample(range(len(train_sentences)), 2600)
    val_sentences = [train_sentences[idx] for idx in val_indices]
    # val_labels = [train_labels[idx] for idx in val_indices]
    val_labels = [0] * len(val_sentences)

    train_sentences = [sample for i, sample in enumerate(train_sentences) if i not in val_indices]
    train_labels = [label for i, label in enumerate(train_labels) if i not in val_indices]

    # Extract prompts with "discrimination, stereotype, and injustice" category as True
    test_filtered_prompts = [test_prompts[i] for i in range(len(test_categories)) if test_categories[i]['non_violent_unethical_behavior']]
    test_filtered_responses = [test_response[i] for i in range(len(test_categories)) if test_categories[i]['non_violent_unethical_behavior']]
    test_filtered_prompt_final = []
    for i in range(len(test_filtered_prompts)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {test_filtered_prompts[i]}\nResponse: {test_filtered_responses[i]}"""
        test_filtered_prompt_final.append(formatted_entry)
    test_filtered_labels = [1] * len(test_filtered_prompt_final)
    test_safe_prompts = [test_prompts[i] for i in range(len(test_prompts)) if test_is_safe[i]][0:7000]
    test_safe_responses = [test_response[i] for i in range(len(test_response)) if test_is_safe[i]][0:7000]
    test_safe_prompt_final = []
    for i in range(len(test_safe_prompts)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {test_safe_prompts[i]}\nResponse: {test_safe_responses[i]}"""
        test_safe_prompt_final.append(formatted_entry)
    test_safe_labels = [0] * len(test_safe_prompt_final)
    test_sentences = test_filtered_prompt_final + test_safe_prompt_final
    test_labels = test_filtered_labels + test_safe_labels
    test_sentences, test_labels = balance_binary_classes(test_sentences, test_labels)
    test_sentences, test_labels = generate_random_data(test_sentences, test_labels, 2600)
    test_labels = [0] * len(test_sentences)

    print('\n\ntrain_labels_IID', len(train_labels))
    print('test_labels_IID', len(test_labels))
    print('val_labels_IID', len(val_labels))
    print('test_labels_IID', test_labels[0:10])
    print('Unique IID train labels', {num: train_labels.count(num) for num in train_labels})
    print('Unique IID test labels', {num: test_labels.count(num) for num in test_labels})
    print('Unique IID val labels', {num: val_labels.count(num) for num in val_labels})
    
    ###Load generated-OOD dataset
    OOD_dataset = load_dataset("openai/gsm8k", 'main')
    # OOD_train_sentences = OOD_dataset['train']['question']
    # OOD_train_responses = OOD_dataset['train']['answer']

    OOD_train_sentences, OOD_train_responses = read_excel_file_responses("data/data_ethical/OOD_samples_llama-3-70b-instruct_math_augmented_beaver-withresponses.xlsx")    
    OOD_train_sentences = OOD_train_sentences[1:]
    OOD_train_responses = OOD_train_responses[1:]

    OOD_train_sentences_final = []
    for i in range(len(OOD_train_sentences)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {OOD_train_sentences[i]}\nResponse: {OOD_train_responses[i]}"""
        OOD_train_sentences_final.append(formatted_entry)
    OOD_train_sentences = OOD_train_sentences_final

    sample_size = 1300
    OOD_val_sentences = random.sample(OOD_train_sentences, sample_size)
    OOD_val_labels = [1] * len(OOD_val_sentences)
    for item in OOD_val_sentences:
        OOD_train_sentences.remove(item)
    OOD_train_labels = [1] * len(OOD_train_sentences)

    OOD_test_sentences = OOD_dataset['test']['question']
    OOD_test_responses = OOD_dataset['test']['answer']
    OOD_test_sentences_final = []
    for i in range(len(OOD_test_sentences)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {OOD_test_sentences[i]}\nResponse: {OOD_test_responses[i]}"""
        OOD_test_sentences_final.append(formatted_entry)
    OOD_test_sentences = OOD_test_sentences_final
    OOD_test_labels = [1] * len(OOD_test_sentences)

    print('\n\nOOD_train_sentences', len(OOD_train_sentences))
    print('OOD_train_labels', len(OOD_train_labels))
    print('OOD_test_sentences', len(OOD_test_sentences))
    print('OOD_test_labels', len(OOD_test_labels))
    print('OOD_val_sentences', len(OOD_val_sentences))
    print('OOD_val_labels', len(OOD_val_labels))

    ###Merge IID+generated-OOD dataset
    train_sentences.extend(OOD_train_sentences)
    train_labels.extend(OOD_train_labels)
    test_sentences.extend(OOD_test_sentences)
    test_labels.extend(OOD_test_labels)
    # test_sentences = OOD_test_sentences
    # test_labels= OOD_test_labels
    val_sentences.extend(OOD_val_sentences)
    val_labels.extend(OOD_val_labels)
    
    print('\n\ntrain_labels_IID+OOD', len(train_labels))
    print('Unique train labels', {num: train_labels.count(num) for num in train_labels})
    print('Unique test labels', {num: test_labels.count(num) for num in test_labels})
    print('Unique val labels', {num: val_labels.count(num) for num in val_labels})
    print('test_labels', len(test_labels))
    print('val_labels', len(val_labels))

    return train_sentences, train_labels, test_sentences, test_labels, val_sentences, val_labels


def load_beavertails_unethical_OOD_mbpp():

    def generate_random_data(samples, labels, num_samples):
        random.seed(42)
        # Create a permutation
        permutation = list(range(len(samples)))
        random.shuffle(permutation)
        
        # Select samples and labels based on the permutation and num_samples
        random_samples = [samples[i] for i in permutation[:num_samples]]
        random_labels = [labels[i] for i in permutation[:num_samples]]
        return random_samples, random_labels

    def balance_binary_classes(samples, labels):
        class_0_samples = [sample for sample, label in zip(samples, labels) if label == 0]
        class_1_samples = [sample for sample, label in zip(samples, labels) if label == 1]
        
        min_samples = min(len(class_0_samples), len(class_1_samples))
        
        balanced_samples = class_0_samples[:min_samples] + class_1_samples[:min_samples]
        balanced_labels = [0] * min_samples + [1] * min_samples
        
        return balanced_samples, balanced_labels

    # Load the training dataset
    train_data = load_dataset('PKU-Alignment/BeaverTails', split='330k_train')
    test_data = load_dataset('PKU-Alignment/BeaverTails', split='330k_test')

    train_prompts = train_data['prompt']
    train_response = train_data['response']
    train_categories = train_data['category']
    train_is_safe = train_data['is_safe']
    test_prompts = test_data['prompt']
    test_response = test_data['response']
    test_categories = test_data['category']
    test_is_safe = test_data['is_safe']

    # Extract prompts with "discrimination, stereotype, and injustice" category as True
    train_filtered_prompts = [train_prompts[i] for i in range(len(train_categories)) if train_categories[i]['non_violent_unethical_behavior']]
    train_filtered_responses = [train_response[i] for i in range(len(train_categories)) if train_categories[i]['non_violent_unethical_behavior']]
    train_filtered_prompt_final = []
    for i in range(len(train_filtered_prompts)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {train_filtered_prompts[i]}\nResponse: {train_filtered_responses[i]}"""
        train_filtered_prompt_final.append(formatted_entry)
    train_filtered_labels = [1] * len(train_filtered_prompt_final)
    train_safe_prompts = [train_prompts[i] for i in range(len(train_prompts)) if train_is_safe[i]][0:60000]
    train_safe_responses = [train_response[i] for i in range(len(train_response)) if train_is_safe[i]][0:60000]
    train_safe_prompt_final = []
    for i in range(len(train_safe_prompts)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {train_safe_prompts[i]}\nResponse: {train_safe_responses[i]}"""
        train_safe_prompt_final.append(formatted_entry)
    train_safe_labels = [0] * len(train_safe_prompt_final)
    train_sentences = train_filtered_prompt_final + train_safe_prompt_final
    train_labels = train_filtered_labels + train_safe_labels
    train_sentences, train_labels = balance_binary_classes(train_sentences, train_labels)
    # for synthetic
    train_sentences, train_labels = generate_random_data(train_sentences, train_labels, 17600)
    train_labels = [0] * len(train_sentences)

    #for synthetic
    val_indices = random.sample(range(len(train_sentences)), 2600)
    val_sentences = [train_sentences[idx] for idx in val_indices]
    # val_labels = [train_labels[idx] for idx in val_indices]
    val_labels = [0] * len(val_sentences)

    train_sentences = [sample for i, sample in enumerate(train_sentences) if i not in val_indices]
    train_labels = [label for i, label in enumerate(train_labels) if i not in val_indices]

    
    # Extract prompts with "discrimination, stereotype, and injustice" category as True
    test_filtered_prompts = [test_prompts[i] for i in range(len(test_categories)) if test_categories[i]['non_violent_unethical_behavior']]
    test_filtered_responses = [test_response[i] for i in range(len(test_categories)) if test_categories[i]['non_violent_unethical_behavior']]
    test_filtered_prompt_final = []
    for i in range(len(test_filtered_prompts)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {test_filtered_prompts[i]}\nResponse: {test_filtered_responses[i]}"""
        test_filtered_prompt_final.append(formatted_entry)
    test_filtered_labels = [1] * len(test_filtered_prompt_final)
    test_safe_prompts = [test_prompts[i] for i in range(len(test_prompts)) if test_is_safe[i]][0:7000]
    test_safe_responses = [test_response[i] for i in range(len(test_response)) if test_is_safe[i]][0:7000]
    test_safe_prompt_final = []
    for i in range(len(test_safe_prompts)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {test_safe_prompts[i]}\nResponse: {test_safe_responses[i]}"""
        test_safe_prompt_final.append(formatted_entry)
    test_safe_labels = [0] * len(test_safe_prompt_final)
    test_sentences = test_filtered_prompt_final + test_safe_prompt_final
    test_labels = test_filtered_labels + test_safe_labels
    test_sentences, test_labels = balance_binary_classes(test_sentences, test_labels)
    test_sentences, test_labels = generate_random_data(test_sentences, test_labels, 1000)
    test_labels = [0] * len(test_sentences)

    print('\n\ntrain_labels_IID', len(train_labels))
    print('test_labels_IID', len(test_labels))
    print('val_labels_IID', len(val_labels))
    print('test_labels_IID', test_labels[0:10])
    print('Unique IID train labels', {num: train_labels.count(num) for num in train_labels})
    print('Unique IID test labels', {num: test_labels.count(num) for num in test_labels})
    print('Unique IID val labels', {num: val_labels.count(num) for num in val_labels})


    ###Load generated-OOD dataset
    OOD_dataset = load_dataset("google-research-datasets/mbpp")
    # OOD_train_sentences = OOD_dataset['train']['text']
    # OOD_train_responses = OOD_dataset['train']['code']

    OOD_train_sentences, OOD_train_responses = read_excel_file_responses("data/data_ethical/OOD_samples_llama-3-70b-instruct_code_augmented_beaver-withresponses.xlsx")    
    OOD_train_sentences = OOD_train_sentences[1:]
    OOD_train_responses = OOD_train_responses[1:]

    OOD_train_sentences_final = []
    for i in range(len(OOD_train_sentences)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {OOD_train_sentences[i]}\nResponse: {OOD_train_responses[i]}"""
        OOD_train_sentences_final.append(formatted_entry)
    OOD_train_sentences = OOD_train_sentences_final

    # OOD_val_sentences = OOD_dataset['validation']['text']
    # OOD_val_responses = OOD_dataset['validation']['code']
    # OOD_val_sentences_final = []
    # for i in range(len(OOD_val_sentences)):
    #     formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {OOD_val_sentences[i]}\nResponse: {OOD_val_responses[i]}"""
    #     OOD_val_sentences_final.append(formatted_entry)
    # OOD_val_sentences = OOD_val_sentences_final

    sample_size = 1300
    OOD_val_sentences = random.sample(OOD_train_sentences, sample_size)
    OOD_val_labels = [1] * len(OOD_val_sentences)
    for item in OOD_val_sentences:
        OOD_train_sentences.remove(item)
    OOD_train_labels = [1] * len(OOD_train_sentences)

    OOD_test_sentences = OOD_dataset['test']['text']
    OOD_test_responses = OOD_dataset['test']['code']
    OOD_test_sentences_final = []
    for i in range(len(OOD_test_sentences)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {OOD_test_sentences[i]}\nResponse: {OOD_test_responses[i]}"""
        OOD_test_sentences_final.append(formatted_entry)
    OOD_test_sentences = OOD_test_sentences_final
    OOD_test_labels = [1] * len(OOD_test_sentences)

    print('\n\nOOD_train_sentences', len(OOD_train_sentences))
    print('OOD_train_labels', len(OOD_train_labels))
    print('OOD_test_sentences', len(OOD_test_sentences))
    print('OOD_test_labels', len(OOD_test_labels))
    print('OOD_val_sentences', len(OOD_val_sentences))
    print('OOD_val_labels', len(OOD_val_labels))


    ###Merge IID+generated-OOD dataset
    train_sentences.extend(OOD_train_sentences)
    train_labels.extend(OOD_train_labels)
    test_sentences.extend(OOD_test_sentences)
    test_labels.extend(OOD_test_labels)
    # test_sentences = OOD_test_sentences
    # test_labels= OOD_test_labels
    val_sentences.extend(OOD_val_sentences)
    val_labels.extend(OOD_val_labels)
    
    print('\n\ntrain_labels_IID+OOD', len(train_labels))
    print('Unique train labels', {num: train_labels.count(num) for num in train_labels})
    print('Unique test labels', {num: test_labels.count(num) for num in test_labels})
    print('Unique val labels', {num: val_labels.count(num) for num in val_labels})
    print('test_labels', len(test_labels))
    print('val_labels', len(val_labels))

    return train_sentences, train_labels, test_sentences, test_labels, val_sentences, val_labels


def load_beavertails_unethical_OOD_sexual_drug():

    def generate_random_data(samples, labels, num_samples):
        random.seed(42)
        # Create a permutation
        permutation = list(range(len(samples)))
        random.shuffle(permutation)
        
        # Select samples and labels based on the permutation and num_samples
        random_samples = [samples[i] for i in permutation[:num_samples]]
        random_labels = [labels[i] for i in permutation[:num_samples]]
        return random_samples, random_labels

    def balance_binary_classes(samples, labels):
        class_0_samples = [sample for sample, label in zip(samples, labels) if label == 0]
        class_1_samples = [sample for sample, label in zip(samples, labels) if label == 1]
        
        min_samples = min(len(class_0_samples), len(class_1_samples))
        
        balanced_samples = class_0_samples[:min_samples] + class_1_samples[:min_samples]
        balanced_labels = [0] * min_samples + [1] * min_samples
        
        return balanced_samples, balanced_labels

    # Load the training dataset
    train_data = load_dataset('PKU-Alignment/BeaverTails', split='330k_train')
    test_data = load_dataset('PKU-Alignment/BeaverTails', split='330k_test')

    train_prompts = train_data['prompt']
    train_response = train_data['response']
    train_categories = train_data['category']
    train_is_safe = train_data['is_safe']
    test_prompts = test_data['prompt']
    test_response = test_data['response']
    test_categories = test_data['category']
    test_is_safe = test_data['is_safe']

    # Extract prompts with "discrimination, stereotype, and injustice" category as True
    train_filtered_prompts = [train_prompts[i] for i in range(len(train_categories)) if train_categories[i]['non_violent_unethical_behavior']]
    train_filtered_responses = [train_response[i] for i in range(len(train_categories)) if train_categories[i]['non_violent_unethical_behavior']]
    train_filtered_prompt_final = []
    for i in range(len(train_filtered_prompts)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {train_filtered_prompts[i]}\nResponse: {train_filtered_responses[i]}"""
        train_filtered_prompt_final.append(formatted_entry)
    train_filtered_labels = [1] * len(train_filtered_prompt_final)
    train_safe_prompts = [train_prompts[i] for i in range(len(train_prompts)) if train_is_safe[i]][0:60000]
    train_safe_responses = [train_response[i] for i in range(len(train_response)) if train_is_safe[i]][0:60000]
    train_safe_prompt_final = []
    for i in range(len(train_safe_prompts)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {train_safe_prompts[i]}\nResponse: {train_safe_responses[i]}"""
        train_safe_prompt_final.append(formatted_entry)
    train_safe_labels = [0] * len(train_safe_prompt_final)
    train_sentences = train_filtered_prompt_final + train_safe_prompt_final
    train_labels = train_filtered_labels + train_safe_labels
    train_sentences, train_labels = balance_binary_classes(train_sentences, train_labels)
    # for synthetic
    train_sentences, train_labels = generate_random_data(train_sentences, train_labels, 17600)
    train_labels = [0] * len(train_sentences)

    #for synthetic
    val_indices = random.sample(range(len(train_sentences)), 2600)
    val_sentences = [train_sentences[idx] for idx in val_indices]
    # val_labels = [train_labels[idx] for idx in val_indices]
    val_labels = [0] * len(val_sentences)

    train_sentences = [sample for i, sample in enumerate(train_sentences) if i not in val_indices]
    train_labels = [label for i, label in enumerate(train_labels) if i not in val_indices]

    
    # Extract prompts with "discrimination, stereotype, and injustice" category as True
    test_filtered_prompts = [test_prompts[i] for i in range(len(test_categories)) if test_categories[i]['non_violent_unethical_behavior']]
    test_filtered_responses = [test_response[i] for i in range(len(test_categories)) if test_categories[i]['non_violent_unethical_behavior']]
    test_filtered_prompt_final = []
    for i in range(len(test_filtered_prompts)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {test_filtered_prompts[i]}\nResponse: {test_filtered_responses[i]}"""
        test_filtered_prompt_final.append(formatted_entry)
    test_filtered_labels = [1] * len(test_filtered_prompt_final)
    test_safe_prompts = [test_prompts[i] for i in range(len(test_prompts)) if test_is_safe[i]][0:7000]
    test_safe_responses = [test_response[i] for i in range(len(test_response)) if test_is_safe[i]][0:7000]
    test_safe_prompt_final = []
    for i in range(len(test_safe_prompts)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {test_safe_prompts[i]}\nResponse: {test_safe_responses[i]}"""
        test_safe_prompt_final.append(formatted_entry)
    test_safe_labels = [0] * len(test_safe_prompt_final)
    test_sentences = test_filtered_prompt_final + test_safe_prompt_final
    test_labels = test_filtered_labels + test_safe_labels
    test_sentences, test_labels = balance_binary_classes(test_sentences, test_labels)
    test_sentences, test_labels = generate_random_data(test_sentences, test_labels, 4600)
    test_labels = [0] * len(test_sentences)

    print('\n\ntrain_labels_IID', len(train_labels))
    print('test_labels_IID', len(test_labels))
    print('val_labels_IID', len(val_labels))
    print('test_labels_IID', test_labels[0:10])
    print('Unique IID train labels', {num: train_labels.count(num) for num in train_labels})
    print('Unique IID test labels', {num: test_labels.count(num) for num in test_labels})
    print('Unique IID val labels', {num: val_labels.count(num) for num in val_labels})


    ###Load generated-OOD dataset containing 'Sexually Explicit, Adult Content' and 'Drug Abuse, Weapons, Banned Substance' categories
    # Train: Extract prompts with 'Sexually Explicit, Adult Content' categories as True
    # OOD_train_filtered_prompts = [train_prompts[i] 
    # for i in range(len(train_categories)) 
    # if (not train_categories[i]['non_violent_unethical_behavior'] 
    #     and train_categories[i]['sexually_explicit,adult_content'] 
    #     )]

    # OOD_train_filtered_responses = [train_response[i] 
    # for i in range(len(train_categories)) 
    # if (not train_categories[i]['non_violent_unethical_behavior'] 
    #     and train_categories[i]['sexually_explicit,adult_content'] 
    #       )]

    OOD_train_filtered_prompts, OOD_train_filtered_responses = read_excel_file_responses("data/data_ethical/OOD_samples_llama-3-70b-instruct_sexual_drug_augmented_beaver-withresponses.xlsx")    
    OOD_train_filtered_prompts = OOD_train_filtered_prompts[1:]
    OOD_train_filtered_responses = OOD_train_filtered_responses[1:]
    filtered_prompts, filtered_responses = zip(*[
    (prompt, response) for prompt, response in zip(OOD_train_filtered_prompts, OOD_train_filtered_responses)
    if 'i cannot' not in response.lower() and 'i can\'t' not in response.lower()])
    OOD_train_filtered_prompts = list(filtered_prompts)
    OOD_train_filtered_responses = list(filtered_responses)

    OOD_train_filtered_prompt_final = []
    for i in range(len(OOD_train_filtered_prompts)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {OOD_train_filtered_prompts[i]}\nResponse: {OOD_train_filtered_responses[i]}"""
        OOD_train_filtered_prompt_final.append(formatted_entry)
    OOD_train_filtered_labels = [0] * len(OOD_train_filtered_prompt_final)

    # Train: Extract prompts with 'Drug Abuse, Weapons, Banned Substance' categories as True
    # OOD_train_filtered_prompts2 = [train_prompts[i] 
    # for i in range(len(train_categories)) 
    # if (not train_categories[i]['non_violent_unethical_behavior'] 
    #     and train_categories[i]['drug_abuse,weapons,banned_substance'] 
    #      )]

    # OOD_train_filtered_responses2 = [train_response[i] 
    # for i in range(len(train_categories)) 
    # if (not train_categories[i]['non_violent_unethical_behavior'] 
    #     and train_categories[i]['drug_abuse,weapons,banned_substance'] 
    #      )]

    OOD_train_filtered_prompts2, OOD_train_filtered_responses2 = read_excel_file_responses("data/data_ethical/OOD_samples_llama-3-70b-instruct_drugs_augmented_beaver-withresponses.xlsx")    
    OOD_train_filtered_prompts2 = OOD_train_filtered_prompts2[1:]
    OOD_train_filtered_responses2 = OOD_train_filtered_responses2[1:]
    
    OOD_train_filtered_prompt_final2 = []
    for i in range(len(OOD_train_filtered_prompts2)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {OOD_train_filtered_prompts2[i]}\nResponse: {OOD_train_filtered_responses2[i]}"""
        OOD_train_filtered_prompt_final2.append(formatted_entry)
    OOD_train_filtered_labels2 = [1] * len(OOD_train_filtered_prompt_final2)

    OOD_train_sentences = OOD_train_filtered_prompt_final + OOD_train_filtered_prompt_final2
    OOD_train_labels = OOD_train_filtered_labels + OOD_train_filtered_labels2
    OOD_train_sentences, OOD_train_labels = balance_binary_classes(OOD_train_sentences, OOD_train_labels)
    OOD_train_sentences, OOD_train_labels = generate_random_data(OOD_train_sentences, OOD_train_labels, 8300)
    # OOD_train_labels = [2] * len(OOD_train_sentences)

    sample_size = 1300
    OOD_val_sentences = random.sample(OOD_train_sentences, sample_size)
    OOD_val_labels = [1] * len(OOD_val_sentences)
    for item in OOD_val_sentences:
        OOD_train_sentences.remove(item)
    OOD_train_labels = [1] * len(OOD_train_sentences)


    # Test: Extract prompts with 'Sexually Explicit, Adult Content' categories as True
    OOD_test_filtered_prompts = [test_prompts[i] 
    for i in range(len(test_categories)) 
    if (not test_categories[i]['non_violent_unethical_behavior'] 
        and test_categories[i]['sexually_explicit,adult_content'] 
         )]

    OOD_test_filtered_responses = [test_response[i] 
    for i in range(len(test_categories)) 
    if (not test_categories[i]['non_violent_unethical_behavior'] 
        and test_categories[i]['sexually_explicit,adult_content'] 
         )]

    OOD_test_filtered_prompt_final = []
    for i in range(len(OOD_test_filtered_prompts)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {OOD_test_filtered_prompts[i]}\nResponse: {OOD_test_filtered_responses[i]}"""
        OOD_test_filtered_prompt_final.append(formatted_entry)
    OOD_test_filtered_labels = [0] * len(OOD_test_filtered_prompt_final)

    # Test: Extract prompts with 'Drug Abuse, Weapons, Banned Substance' categories as True
    OOD_test_filtered_prompts2 = [test_prompts[i] 
    for i in range(len(test_categories)) 
    if (not test_categories[i]['non_violent_unethical_behavior'] 
        and test_categories[i]['drug_abuse,weapons,banned_substance'] 
         )]

    OOD_test_filtered_responses2 = [test_response[i] 
    for i in range(len(test_categories)) 
    if (not test_categories[i]['non_violent_unethical_behavior'] 
        and test_categories[i]['drug_abuse,weapons,banned_substance'] 
         )]

    OOD_test_filtered_prompt_final2 = []
    for i in range(len(OOD_test_filtered_prompts2)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {OOD_test_filtered_prompts2[i]}\nResponse: {OOD_test_filtered_responses2[i]}"""
        OOD_test_filtered_prompt_final2.append(formatted_entry)
    OOD_test_filtered_labels2 = [1] * len(OOD_test_filtered_prompt_final2)

    OOD_test_sentences = OOD_test_filtered_prompt_final + OOD_test_filtered_prompt_final2
    OOD_test_labels = OOD_test_filtered_labels + OOD_test_filtered_labels2
    # OOD_test_sentences, OOD_test_labels = balance_binary_classes(OOD_test_sentences, OOD_test_labels)
    # OOD_test_sentences, OOD_test_labels = generate_random_data(OOD_test_sentences, OOD_test_labels, 2000)
    OOD_test_labels = [1] * len(OOD_test_sentences)

    print('\n\nOOD_train_sentences', len(OOD_train_sentences))
    print('OOD_train_labels', len(OOD_train_labels))
    print('OOD_test_sentences', len(OOD_test_sentences))
    print('OOD_test_labels', len(OOD_test_labels))
    print('OOD_val_sentences', len(OOD_val_sentences))
    print('OOD_val_labels', len(OOD_val_labels))


    ###Merge IID+generated-OOD dataset
    train_sentences.extend(OOD_train_sentences)
    train_labels.extend(OOD_train_labels)
    test_sentences.extend(OOD_test_sentences)
    test_labels.extend(OOD_test_labels)
    # test_sentences = OOD_test_sentences
    # test_labels= OOD_test_labels
    val_sentences.extend(OOD_val_sentences)
    val_labels.extend(OOD_val_labels)
    
    print('\n\ntrain_labels_IID+OOD', len(train_labels))
    print('Unique train labels', {num: train_labels.count(num) for num in train_labels})
    print('Unique test labels', {num: test_labels.count(num) for num in test_labels})
    print('Unique val labels', {num: val_labels.count(num) for num in val_labels})
    print('test_labels', len(test_labels))
    print('val_labels', len(val_labels))

    return train_sentences, train_labels, test_sentences, test_labels, val_sentences, val_labels


def load_beavertails_unethical_OOD_discrimincation_hate():

    def generate_random_data(samples, labels, num_samples):
        random.seed(42)
        # Create a permutation
        permutation = list(range(len(samples)))
        random.shuffle(permutation)
        
        # Select samples and labels based on the permutation and num_samples
        random_samples = [samples[i] for i in permutation[:num_samples]]
        random_labels = [labels[i] for i in permutation[:num_samples]]
        return random_samples, random_labels

    def balance_binary_classes(samples, labels):
        class_0_samples = [sample for sample, label in zip(samples, labels) if label == 0]
        class_1_samples = [sample for sample, label in zip(samples, labels) if label == 1]
        
        min_samples = min(len(class_0_samples), len(class_1_samples))
        
        balanced_samples = class_0_samples[:min_samples] + class_1_samples[:min_samples]
        balanced_labels = [0] * min_samples + [1] * min_samples
        
        return balanced_samples, balanced_labels

    # Load the training dataset
    train_data = load_dataset('PKU-Alignment/BeaverTails', split='330k_train')
    test_data = load_dataset('PKU-Alignment/BeaverTails', split='330k_test')

    train_prompts = train_data['prompt']
    train_response = train_data['response']
    train_categories = train_data['category']
    train_is_safe = train_data['is_safe']
    test_prompts = test_data['prompt']
    test_response = test_data['response']
    test_categories = test_data['category']
    test_is_safe = test_data['is_safe']

    # Extract prompts with "discrimination, stereotype, and injustice" category as True
    # train_filtered_prompts = [train_prompts[i] for i in range(len(train_categories)) if train_categories[i]['non_violent_unethical_behavior']]
    # train_filtered_responses = [train_response[i] for i in range(len(train_categories)) if train_categories[i]['non_violent_unethical_behavior']]
    train_filtered_prompts = [train_prompts[i] 
    for i in range(len(train_categories)) 
    if (train_categories[i]['non_violent_unethical_behavior'] 
        and not train_categories[i]['discrimination,stereotype,injustice'] 
        and not train_categories[i]['hate_speech,offensive_language'] 
        )]

    train_filtered_responses = [train_response[i] 
    for i in range(len(train_categories)) 
    if (train_categories[i]['non_violent_unethical_behavior'] 
        and not train_categories[i]['discrimination,stereotype,injustice'] 
        and not train_categories[i]['hate_speech,offensive_language'] 
        )]

    train_filtered_prompt_final = []
    for i in range(len(train_filtered_prompts)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {train_filtered_prompts[i]}\nResponse: {train_filtered_responses[i]}"""
        train_filtered_prompt_final.append(formatted_entry)
    train_filtered_labels = [1] * len(train_filtered_prompt_final)
    train_safe_prompts = [train_prompts[i] for i in range(len(train_prompts)) if train_is_safe[i]][0:60000]
    train_safe_responses = [train_response[i] for i in range(len(train_response)) if train_is_safe[i]][0:60000]

    train_safe_prompt_final = []
    for i in range(len(train_safe_prompts)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {train_safe_prompts[i]}\nResponse: {train_safe_responses[i]}"""
        train_safe_prompt_final.append(formatted_entry)
    train_safe_labels = [0] * len(train_safe_prompt_final)
    train_sentences = train_filtered_prompt_final + train_safe_prompt_final
    train_labels = train_filtered_labels + train_safe_labels
    train_sentences, train_labels = balance_binary_classes(train_sentences, train_labels)
    # for synthetic
    train_sentences, train_labels = generate_random_data(train_sentences, train_labels, 13000)
    train_labels = [0] * len(train_sentences)


    #for synthetic
    val_indices = random.sample(range(len(train_sentences)), 2600)
    val_sentences = [train_sentences[idx] for idx in val_indices]
    # val_labels = [train_labels[idx] for idx in val_indices]
    val_labels = [0] * len(val_sentences)

    train_sentences = [sample for i, sample in enumerate(train_sentences) if i not in val_indices]
    train_labels = [label for i, label in enumerate(train_labels) if i not in val_indices]

    
    # Extract prompts with "non_violent_unethical_behavior" category as True
    # test_filtered_prompts = [test_prompts[i] for i in range(len(test_categories)) if test_categories[i]['non_violent_unethical_behavior']]
    # test_filtered_responses = [test_response[i] for i in range(len(test_categories)) if test_categories[i]['non_violent_unethical_behavior']]
    test_filtered_prompts = [test_prompts[i] 
        for i in range(len(test_categories)) 
        if (test_categories[i]['non_violent_unethical_behavior'] 
            and not test_categories[i]['discrimination,stereotype,injustice'] 
            and not test_categories[i]['hate_speech,offensive_language'] 
            )]

    test_filtered_responses = [test_response[i] 
    for i in range(len(test_categories)) 
    if (test_categories[i]['non_violent_unethical_behavior'] 
        and not test_categories[i]['discrimination,stereotype,injustice'] 
        and not test_categories[i]['hate_speech,offensive_language'] 
        )]


    test_filtered_prompt_final = []
    for i in range(len(test_filtered_prompts)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {test_filtered_prompts[i]}\nResponse: {test_filtered_responses[i]}"""
        test_filtered_prompt_final.append(formatted_entry)
    test_filtered_labels = [1] * len(test_filtered_prompt_final)
    test_safe_prompts = [test_prompts[i] for i in range(len(test_prompts)) if test_is_safe[i]][0:7000]
    test_safe_responses = [test_response[i] for i in range(len(test_response)) if test_is_safe[i]][0:7000]

    test_safe_prompt_final = []
    for i in range(len(test_safe_prompts)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {test_safe_prompts[i]}\nResponse: {test_safe_responses[i]}"""
        test_safe_prompt_final.append(formatted_entry)
    test_safe_labels = [0] * len(test_safe_prompt_final)
    test_sentences = test_filtered_prompt_final + test_safe_prompt_final
    test_labels = test_filtered_labels + test_safe_labels
    test_sentences, test_labels = balance_binary_classes(test_sentences, test_labels)
    test_sentences, test_labels = generate_random_data(test_sentences, test_labels, 1400)
    test_labels = [0] * len(test_sentences)

    print('\n\ntrain_labels_IID', len(train_labels))
    print('test_labels_IID', len(test_labels))
    print('val_labels_IID', len(val_labels))
    print('test_labels_IID', test_labels[0:10])
    print('Unique IID train labels', {num: train_labels.count(num) for num in train_labels})
    print('Unique IID test labels', {num: test_labels.count(num) for num in test_labels})
    print('Unique IID val labels', {num: val_labels.count(num) for num in val_labels})


    ###Load generated-OOD dataset containing 'Sexually Explicit, Adult Content' and 'Drug Abuse, Weapons, Banned Substance' categories
    # Train: Extract prompts with 'discrimination,stereotype,injustice' categories as True
    # OOD_train_filtered_prompts = [train_prompts[i] 
    # for i in range(len(train_categories)) 
    # if (not train_categories[i]['non_violent_unethical_behavior'] 
    #     and train_categories[i]['discrimination,stereotype,injustice'] 
    #     )]

    # OOD_train_filtered_responses = [train_response[i] 
    # for i in range(len(train_categories)) 
    # if (not train_categories[i]['non_violent_unethical_behavior'] 
    #     and train_categories[i]['discrimination,stereotype,injustice'] 
    #       )]

    OOD_train_filtered_prompts, OOD_train_filtered_responses = read_excel_file_responses("data/data_ethical/OOD_samples_llama-3-70b-instruct_discrim_augmented_beaver-withresponses.xlsx")    
    OOD_train_filtered_prompts = OOD_train_filtered_prompts[1:]
    OOD_train_filtered_responses = OOD_train_filtered_responses[1:]

    OOD_train_filtered_prompt_final = []
    for i in range(len(OOD_train_filtered_prompts)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {OOD_train_filtered_prompts[i]}\nResponse: {OOD_train_filtered_responses[i]}"""
        OOD_train_filtered_prompt_final.append(formatted_entry)
    OOD_train_filtered_labels = [0] * len(OOD_train_filtered_prompt_final)

    # Train: Extract prompts with 'hate_speech,offensive_language' categories as True
    # OOD_train_filtered_prompts2 = [train_prompts[i] 
    # for i in range(len(train_categories)) 
    # if (not train_categories[i]['non_violent_unethical_behavior'] 
    #     and train_categories[i]['hate_speech,offensive_language'] 
    #      )]

    # OOD_train_filtered_responses2 = [train_response[i] 
    # for i in range(len(train_categories)) 
    # if (not train_categories[i]['non_violent_unethical_behavior'] 
    #     and train_categories[i]['hate_speech,offensive_language'] 
    #      )]

    OOD_train_filtered_prompts2, OOD_train_filtered_responses2 = read_excel_file_responses("data/data_ethical/OOD_samples_llama-3-70b-instruct_hate_augmented_beaver-withresponses.xlsx")    
    OOD_train_filtered_prompts2 = OOD_train_filtered_prompts2[1:]
    OOD_train_filtered_responses2 = OOD_train_filtered_responses2[1:]

    OOD_train_filtered_prompt_final2 = []
    for i in range(len(OOD_train_filtered_prompts2)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {OOD_train_filtered_prompts2[i]}\nResponse: {OOD_train_filtered_responses2[i]}"""
        OOD_train_filtered_prompt_final2.append(formatted_entry)
    OOD_train_filtered_labels2 = [1] * len(OOD_train_filtered_prompt_final2)

    OOD_train_sentences = OOD_train_filtered_prompt_final + OOD_train_filtered_prompt_final2
    OOD_train_labels = OOD_train_filtered_labels + OOD_train_filtered_labels2
    OOD_train_sentences, OOD_train_labels = balance_binary_classes(OOD_train_sentences, OOD_train_labels)
    OOD_train_sentences, OOD_train_labels = generate_random_data(OOD_train_sentences, OOD_train_labels, 6500)
    # OOD_train_labels = [2] * len(OOD_train_sentences)

    sample_size = 1300
    OOD_val_sentences = random.sample(OOD_train_sentences, sample_size)
    OOD_val_labels = [1] * len(OOD_val_sentences)
    for item in OOD_val_sentences:
        OOD_train_sentences.remove(item)
    OOD_train_labels = [1] * len(OOD_train_sentences)


    # Test: Extract prompts with 'discrimination,stereotype,injustice' categories as True
    OOD_test_filtered_prompts = [test_prompts[i] 
    for i in range(len(test_categories)) 
    if (not test_categories[i]['non_violent_unethical_behavior'] 
        and test_categories[i]['discrimination,stereotype,injustice'] 
         )]

    OOD_test_filtered_responses = [test_response[i] 
    for i in range(len(test_categories)) 
    if (not test_categories[i]['non_violent_unethical_behavior'] 
        and test_categories[i]['discrimination,stereotype,injustice'] 
         )]

    OOD_test_filtered_prompt_final = []
    for i in range(len(OOD_test_filtered_prompts)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {OOD_test_filtered_prompts[i]}\nResponse: {OOD_test_filtered_responses[i]}"""
        OOD_test_filtered_prompt_final.append(formatted_entry)
    OOD_test_filtered_labels = [0] * len(OOD_test_filtered_prompt_final)

    # Test: Extract prompts with 'hate_speech,offensive_language' categories as True
    OOD_test_filtered_prompts2 = [test_prompts[i] 
    for i in range(len(test_categories)) 
    if (not test_categories[i]['non_violent_unethical_behavior'] 
        and test_categories[i]['hate_speech,offensive_language'] 
         )]

    OOD_test_filtered_responses2 = [test_response[i] 
    for i in range(len(test_categories)) 
    if (not test_categories[i]['non_violent_unethical_behavior'] 
        and test_categories[i]['hate_speech,offensive_language'] 
         )]

    OOD_test_filtered_prompt_final2 = []
    for i in range(len(OOD_test_filtered_prompts2)):
        formatted_entry = f"""Given a user text as input to an LLM and its response, classify the sentiment of the LLM's response as Positive, Negative, or Neutral.\nText: {OOD_test_filtered_prompts2[i]}\nResponse: {OOD_test_filtered_responses2[i]}"""
        OOD_test_filtered_prompt_final2.append(formatted_entry)
    OOD_test_filtered_labels2 = [1] * len(OOD_test_filtered_prompt_final2)

    OOD_test_sentences = OOD_test_filtered_prompt_final + OOD_test_filtered_prompt_final2
    OOD_test_labels = OOD_test_filtered_labels + OOD_test_filtered_labels2
    # OOD_test_sentences, OOD_test_labels = balance_binary_classes(OOD_test_sentences, OOD_test_labels)
    # OOD_test_sentences, OOD_test_labels = generate_random_data(OOD_test_sentences, OOD_test_labels, 2000)
    OOD_test_labels = [1] * len(OOD_test_sentences)

    print('\n\nOOD_train_sentences', len(OOD_train_sentences))
    print('OOD_train_labels', len(OOD_train_labels))
    print('OOD_test_sentences', len(OOD_test_sentences))
    print('OOD_test_labels', len(OOD_test_labels))
    print('OOD_val_sentences', len(OOD_val_sentences))
    print('OOD_val_labels', len(OOD_val_labels))

    ###Merge IID+generated-OOD dataset
    train_sentences.extend(OOD_train_sentences)
    train_labels.extend(OOD_train_labels)
    test_sentences.extend(OOD_test_sentences)
    test_labels.extend(OOD_test_labels)
    # test_sentences = OOD_test_sentences
    # test_labels= OOD_test_labels
    val_sentences.extend(OOD_val_sentences)
    val_labels.extend(OOD_val_labels)
    
    print('\n\ntrain_labels_IID+OOD', len(train_labels))
    print('Unique train labels', {num: train_labels.count(num) for num in train_labels})
    print('Unique test labels', {num: test_labels.count(num) for num in test_labels})
    print('Unique val labels', {num: val_labels.count(num) for num in val_labels})
    print('test_labels', len(test_labels))
    print('val_labels', len(val_labels))

    return train_sentences, train_labels, test_sentences, test_labels, val_sentences, val_labels


def read_excel_file(file_path):
    strings_list = []
    try:
        # Load the workbook
        workbook = load_workbook(file_path)
        # Select the first sheet
        sheet = workbook.active
        # Iterate over rows in the first column and append strings to the list
        for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row, values_only=True):
            if row[0]:
                strings_list.append(row[0])
    except Exception as e:
        print("An error occurred:", e)
    return strings_list


def read_excel_file_responses(file_path):
    column1_strings = []  # List to store strings from the first column
    column2_strings = []  # List to store strings from the second column
    try:
        # Load the workbook
        workbook = load_workbook(file_path)
        # Select the first sheet
        sheet = workbook.active
        # Iterate over rows in the first column and append strings to the list
        for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row, values_only=True):
            column1_strings.append(row[0])
        # Iterate over rows in the second column and append strings to the list
        for row in sheet.iter_rows(min_row=1, min_col=2, max_col=2, max_row=sheet.max_row, values_only=True):
            column2_strings.append(row[0])
    except Exception as e:
        print("An error occurred:", e)
    return column1_strings, column2_strings


def load_civil_comments_toxicity_OOD_gsm8k():

    def generate_random_data(samples, labels, num_samples):
        # Set the seed for reproducibility
        random.seed(42)

        # Create a permutation
        permutation = list(range(len(samples)))
        random.shuffle(permutation)
        
        # Select samples and labels based on the permutation and num_samples
        random_samples = [samples[i] for i in permutation[:num_samples]]
        random_labels = [labels[i] for i in permutation[:num_samples]]
        return random_samples, random_labels
        
    def convert_labels(labels):
    # return [1 if label > 0.6 else 0 if 0.1 <= label <= 0.4 else None for label in labels if label is not None]
        return [1 if label > 0.6 else 0 if label == 0.0 else None for label in labels if label is not None]

    def get_training_samples(samples, labels):
        return [sample for sample, label in zip(samples, labels) if label is not None]


    def balance_binary_classes(samples, labels):
        class_0_samples = [sample for sample, label in zip(samples, labels) if label == 0]
        class_1_samples = [sample for sample, label in zip(samples, labels) if label == 1]
        
        min_samples = min(len(class_0_samples), len(class_1_samples))
        
        balanced_samples = class_0_samples[:min_samples] + class_1_samples[:min_samples]
        balanced_labels = [0] * min_samples + [1] * min_samples
        
        return balanced_samples, balanced_labels

    dataset = load_dataset("google/civil_comments")

    train_labels =  convert_labels(dataset['train']['toxicity'])
    train_sentences = get_training_samples(dataset['train']['text'], train_labels)
    train_labels = [entry for entry in train_labels if entry is not None]
    train_sentences, train_labels = balance_binary_classes(train_sentences, train_labels)
    train_sentences, train_labels = generate_random_data(train_sentences, train_labels, 15000)
    train_labels = [0] * len(train_sentences)

    test_labels =  convert_labels(dataset['test']['toxicity'])
    test_sentences = get_training_samples(dataset['test']['text'], test_labels)
    test_labels = [entry for entry in test_labels if entry is not None]
    test_sentences, test_labels = balance_binary_classes(test_sentences, test_labels)
    test_sentences, test_labels = generate_random_data(test_sentences, test_labels, 2600)
    test_labels = [0] * len(test_sentences)
    
    val_labels =  convert_labels(dataset['validation']['toxicity'])
    val_sentences = get_training_samples(dataset['validation']['text'], val_labels)
    val_labels = [entry for entry in val_labels if entry is not None]
    val_sentences, val_labels = balance_binary_classes(val_sentences, val_labels)
    val_sentences, val_labels = generate_random_data(val_sentences, val_labels, 2600)
    val_labels = [0] * len(val_sentences)

    print('\n\ntrain_labels_IID', len(train_labels))
    print('test_labels_IID', test_labels[0:10])
    print('Unique IID test labels', {num: test_labels.count(num) for num in test_labels})
    print('test_labels_IID', len(test_labels))
    print('val_labels_IID', len(val_labels))

    ###Load generated-OOD dataset
    OOD_dataset = load_dataset("openai/gsm8k", 'main')
    OOD_train_sentences = OOD_dataset['train']['question']
    # OOD_train_labels = OOD_dataset['train']['answer']
    # OOD_train_sentences = read_excel_file("data/data_ethical/OOD_samples_llama-3-70b-instruct_math_augmented.xlsx")
    # OOD_train_sentences = OOD_train_sentences[1:]

    OOD_test_sentences = OOD_dataset['test']['question']
    # OOD_test_labels = OOD_dataset['test']['answer']
    OOD_test_labels = [1] * len(OOD_test_sentences)

    sample_size = min(len(OOD_train_sentences) // 2, 1300)
    OOD_val_sentences = random.sample(OOD_train_sentences, sample_size)
    OOD_val_labels = [1] * len(OOD_val_sentences)
    for item in OOD_val_sentences:
        OOD_train_sentences.remove(item)
    OOD_train_labels = [1] * len(OOD_train_sentences)

    ###Merge IID+generated-OOD dataset
    train_sentences.extend(OOD_train_sentences)
    train_labels.extend(OOD_train_labels)
    test_sentences.extend(OOD_test_sentences)
    test_labels.extend(OOD_test_labels)
    # # test_sentences = OOD_test_sentences
    # # test_labels= OOD_test_labels
    val_sentences.extend(OOD_val_sentences)
    val_labels.extend(OOD_val_labels)
    
    print('\n\ntrain_labels_IID+OOD', len(train_labels))
    print('Unique train labels', {num: train_labels.count(num) for num in train_labels})
    print('Unique test labels', {num: test_labels.count(num) for num in test_labels})
    print('Unique val labels', {num: val_labels.count(num) for num in val_labels})
    print('test_labels', len(test_labels))
    print('val_labels', len(val_labels))

    return train_sentences, train_labels, test_sentences, test_labels, val_sentences, val_labels


def load_civil_comments_toxicity_OOD_mbpp():

    def generate_random_data(samples, labels, num_samples):
        random.seed(42)
        # Create a permutation
        permutation = list(range(len(samples)))
        random.shuffle(permutation)
        
        # Select samples and labels based on the permutation and num_samples
        random_samples = [samples[i] for i in permutation[:num_samples]]
        random_labels = [labels[i] for i in permutation[:num_samples]]
        return random_samples, random_labels
        
    def convert_labels(labels):
    # return [1 if label > 0.6 else 0 if 0.1 <= label <= 0.4 else None for label in labels if label is not None]
        return [1 if label > 0.6 else 0 if label == 0.0 else None for label in labels if label is not None]

    def get_training_samples(samples, labels):
        return [sample for sample, label in zip(samples, labels) if label is not None]


    def balance_binary_classes(samples, labels):
        class_0_samples = [sample for sample, label in zip(samples, labels) if label == 0]
        class_1_samples = [sample for sample, label in zip(samples, labels) if label == 1]
        
        min_samples = min(len(class_0_samples), len(class_1_samples))
        
        balanced_samples = class_0_samples[:min_samples] + class_1_samples[:min_samples]
        balanced_labels = [0] * min_samples + [1] * min_samples
        
        return balanced_samples, balanced_labels

    dataset = load_dataset("google/civil_comments")

    train_labels =  convert_labels(dataset['train']['toxicity'])
    train_sentences = get_training_samples(dataset['train']['text'], train_labels)
    train_labels = [entry for entry in train_labels if entry is not None]
    train_sentences, train_labels = balance_binary_classes(train_sentences, train_labels)
    train_sentences, train_labels = generate_random_data(train_sentences, train_labels, 15000)
    train_labels = [0] * len(train_sentences)

    test_labels =  convert_labels(dataset['test']['toxicity'])
    test_sentences = get_training_samples(dataset['test']['text'], test_labels)
    test_labels = [entry for entry in test_labels if entry is not None]
    test_sentences, test_labels = balance_binary_classes(test_sentences, test_labels)
    test_sentences, test_labels = generate_random_data(test_sentences, test_labels, 1000)
    test_labels = [0] * len(test_sentences)
    
    val_labels =  convert_labels(dataset['validation']['toxicity'])
    val_sentences = get_training_samples(dataset['validation']['text'], val_labels)
    val_labels = [entry for entry in val_labels if entry is not None]
    val_sentences, val_labels = balance_binary_classes(val_sentences, val_labels)
    val_sentences, val_labels = generate_random_data(val_sentences, val_labels, 2600)
    val_labels = [0] * len(val_sentences)

    print('\n\ntrain_labels_IID', len(train_labels))
    print('test_labels_IID', test_labels[0:10])
    print('Unique IID test labels', {num: test_labels.count(num) for num in test_labels})
    print('test_labels_IID', len(test_labels))
    print('val_labels_IID', len(val_labels))

    ###Load generated-OOD dataset
    OOD_dataset = load_dataset("google-research-datasets/mbpp")
    OOD_train_sentences = OOD_dataset['train']['text']
    # OOD_train_labels = OOD_dataset['train']['answer']
    # OOD_train_sentences = read_excel_file("data/data_ethical/OOD_samples_llama-3-70b-instruct_code_augmented.xlsx")
    # OOD_train_sentences = OOD_train_sentences[1:]

    OOD_test_sentences = OOD_dataset['test']['text']
    OOD_test_labels = [1] * len(OOD_test_sentences)

    OOD_val_sentences = OOD_dataset['validation']['text']

    # sample_size = 1300
    # OOD_val_sentences = random.sample(OOD_train_sentences, sample_size)
    OOD_val_labels = [1] * len(OOD_val_sentences)
    # for item in OOD_val_sentences:
    #     OOD_train_sentences.remove(item)
    OOD_train_labels = [1] * len(OOD_train_sentences)

    print('\n\nOOD_train_sentences', len(OOD_train_sentences))
    print('OOD_train_labels', len(OOD_train_labels))
    print('OOD_test_sentences', len(OOD_test_sentences))
    print('OOD_test_labels', len(OOD_test_labels))
    print('OOD_val_sentences', len(OOD_val_sentences))
    print('OOD_val_labels', len(OOD_val_labels))

    ###Merge IID+generated-OOD dataset
    train_sentences.extend(OOD_train_sentences)
    train_labels.extend(OOD_train_labels)
    test_sentences.extend(OOD_test_sentences)
    test_labels.extend(OOD_test_labels)
    # test_sentences = OOD_test_sentences
    # test_labels= OOD_test_labels
    val_sentences.extend(OOD_val_sentences)
    val_labels.extend(OOD_val_labels)
    
    print('\n\ntrain_labels_IID+OOD', len(train_labels))
    print('Unique train labels', {num: train_labels.count(num) for num in train_labels})
    print('Unique test labels', {num: test_labels.count(num) for num in test_labels})
    print('Unique val labels', {num: val_labels.count(num) for num in val_labels})
    print('test_labels', len(test_labels))
    print('val_labels', len(val_labels))

    return train_sentences, train_labels, test_sentences, test_labels, val_sentences, val_labels

def load_civil_comments_toxicity_OOD_sst2():

    def generate_random_data(samples, labels, num_samples):
        random.seed(42)
        # Create a permutation
        permutation = list(range(len(samples)))
        random.shuffle(permutation)
        
        # Select samples and labels based on the permutation and num_samples
        random_samples = [samples[i] for i in permutation[:num_samples]]
        random_labels = [labels[i] for i in permutation[:num_samples]]
        return random_samples, random_labels
        
    def convert_labels(labels):
    # return [1 if label > 0.6 else 0 if 0.1 <= label <= 0.4 else None for label in labels if label is not None]
        return [1 if label > 0.6 else 0 if label == 0.0 else None for label in labels if label is not None]

    def get_training_samples(samples, labels):
        return [sample for sample, label in zip(samples, labels) if label is not None]


    def balance_binary_classes(samples, labels):
        class_0_samples = [sample for sample, label in zip(samples, labels) if label == 0]
        class_1_samples = [sample for sample, label in zip(samples, labels) if label == 1]
        
        min_samples = min(len(class_0_samples), len(class_1_samples))
        
        balanced_samples = class_0_samples[:min_samples] + class_1_samples[:min_samples]
        balanced_labels = [0] * min_samples + [1] * min_samples
        
        return balanced_samples, balanced_labels

    dataset = load_dataset("google/civil_comments")

    train_labels =  convert_labels(dataset['train']['toxicity'])
    train_sentences = get_training_samples(dataset['train']['text'], train_labels)
    train_labels = [entry for entry in train_labels if entry is not None]
    train_sentences, train_labels = balance_binary_classes(train_sentences, train_labels)
    train_sentences, train_labels = generate_random_data(train_sentences, train_labels, 15000)
    train_labels = [0] * len(train_sentences)

    test_labels =  convert_labels(dataset['test']['toxicity'])
    test_sentences = get_training_samples(dataset['test']['text'], test_labels)
    test_labels = [entry for entry in test_labels if entry is not None]
    test_sentences, test_labels = balance_binary_classes(test_sentences, test_labels)
    test_sentences, test_labels = generate_random_data(test_sentences, test_labels, 3650)
    test_labels = [0] * len(test_sentences)
    
    val_labels =  convert_labels(dataset['validation']['toxicity'])
    val_sentences = get_training_samples(dataset['validation']['text'], val_labels)
    val_labels = [entry for entry in val_labels if entry is not None]
    val_sentences, val_labels = balance_binary_classes(val_sentences, val_labels)
    val_sentences, val_labels = generate_random_data(val_sentences, val_labels, 1800)
    val_labels = [0] * len(val_sentences)

    print('\n\ntrain_labels_IID', len(train_labels))
    print('test_labels_IID', test_labels[0:10])
    print('Unique IID test labels', {num: test_labels.count(num) for num in test_labels})
    print('test_labels_IID', len(test_labels))
    print('val_labels_IID', len(val_labels))

    ###Load generated-OOD dataset
    # OOD_train_sentences = read_excel_file("data/data_ethical/OOD_samples_llama-3-70b-instruct_sst2_augmented.xlsx")
    # OOD_train_sentences = OOD_train_sentences[1:]
    # OOD_train_sentences = [review for review in OOD_train_sentences if len(review) >= 15]

    OOD_train_sentences, OOD_train_labels, OOD_test_sentences, OOD_test_labels, OOD_val_sentences, OOD_val_labels= load_sst2()
    OOD_train_labels = [1] * len(OOD_train_sentences)
    OOD_test_labels = [1] * len(OOD_test_sentences)
    OOD_val_labels = [1] * len(OOD_val_sentences)
    
    # sample_size = min(len(OOD_train_sentences) // 2, 1300)
    # OOD_val_sentences = random.sample(OOD_train_sentences, sample_size)
    # OOD_val_labels = [2] * len(OOD_val_sentences)
    # for item in OOD_val_sentences:
    #     OOD_train_sentences.remove(item)
    
    ###Merge IID+generated-OOD dataset
    train_sentences.extend(OOD_train_sentences)
    train_labels.extend(OOD_train_labels)
    test_sentences.extend(OOD_test_sentences)
    test_labels.extend(OOD_test_labels)
    # test_sentences = OOD_test_sentences
    # test_labels= OOD_test_labels
    val_sentences.extend(OOD_val_sentences)
    val_labels.extend(OOD_val_labels)
    
    print('\n\ntrain_labels_IID+OOD', len(train_labels))
    print('Unique train labels', {num: train_labels.count(num) for num in train_labels})
    print('Unique test labels', {num: test_labels.count(num) for num in test_labels})
    print('Unique val labels', {num: val_labels.count(num) for num in val_labels})
    print('test_labels', len(test_labels))
    print('val_labels', len(val_labels))

    return train_sentences, train_labels, test_sentences, test_labels, val_sentences, val_labels


def load_civil_comments_toxicity_OOD_toxigen():

    def generate_random_data(samples, labels, num_samples):
        random.seed(42)
        # Create a permutation
        permutation = list(range(len(samples)))
        random.shuffle(permutation)
        
        # Select samples and labels based on the permutation and num_samples
        random_samples = [samples[i] for i in permutation[:num_samples]]
        random_labels = [labels[i] for i in permutation[:num_samples]]
        return random_samples, random_labels
        
    def convert_labels(labels):
    # return [1 if label > 0.6 else 0 if 0.1 <= label <= 0.4 else None for label in labels if label is not None]
        return [1 if label > 0.6 else 0 if label == 0.0 else None for label in labels if label is not None]

    def get_training_samples(samples, labels):
        return [sample for sample, label in zip(samples, labels) if label is not None]


    def balance_binary_classes(samples, labels):
        class_0_samples = [sample for sample, label in zip(samples, labels) if label == 0]
        class_1_samples = [sample for sample, label in zip(samples, labels) if label == 1]
        
        min_samples = min(len(class_0_samples), len(class_1_samples))
        
        balanced_samples = class_0_samples[:min_samples] + class_1_samples[:min_samples]
        balanced_labels = [0] * min_samples + [1] * min_samples
        
        return balanced_samples, balanced_labels

    dataset = load_dataset("google/civil_comments")

    train_labels =  convert_labels(dataset['train']['toxicity'])
    train_sentences = get_training_samples(dataset['train']['text'], train_labels)
    train_labels = [entry for entry in train_labels if entry is not None]
    train_sentences, train_labels = balance_binary_classes(train_sentences, train_labels)
    train_sentences, train_labels = generate_random_data(train_sentences, train_labels, 14000)
    train_labels = [0] * len(train_sentences)

    test_labels =  convert_labels(dataset['test']['toxicity'])
    test_sentences = get_training_samples(dataset['test']['text'], test_labels)
    test_labels = [entry for entry in test_labels if entry is not None]
    test_sentences, test_labels = balance_binary_classes(test_sentences, test_labels)
    test_sentences, test_labels = generate_random_data(test_sentences, test_labels, 1900)
    test_labels = [0] * len(test_sentences)
    
    val_labels =  convert_labels(dataset['validation']['toxicity'])
    val_sentences = get_training_samples(dataset['validation']['text'], val_labels)
    val_labels = [entry for entry in val_labels if entry is not None]
    val_sentences, val_labels = balance_binary_classes(val_sentences, val_labels)
    val_sentences, val_labels = generate_random_data(val_sentences, val_labels, 4000)
    val_labels = [0] * len(val_sentences)

    print('\n\ntrain_labels_IID', len(train_labels))
    print('test_labels_IID', test_labels[0:10])
    print('Unique IID test labels', {num: test_labels.count(num) for num in test_labels})
    print('test_labels_IID', len(test_labels))
    print('val_labels_IID', len(val_labels))

    ###Load generated-OOD dataset
    OOD_dataset = load_dataset("toxigen/toxigen-data", name="annotated", use_auth_token=True)
    OOD_train_sentences = OOD_dataset['train']['text']
    OOD_test_sentences = OOD_dataset['test']['text']
    
    sample_size = 2000
    OOD_val_sentences = random.sample(OOD_train_sentences, sample_size)
    for item in OOD_val_sentences:
        OOD_train_sentences.remove(item)

    # OOD_train_sentences = read_excel_file("data/data_ethical/OOD_samples_llama-3-70b-instruct_toxigen_augmented.xlsx")
    # OOD_train_sentences = OOD_train_sentences[1:]

    OOD_train_labels = [1] * len(OOD_train_sentences)
    OOD_test_labels = [1] * len(OOD_test_sentences)
    OOD_val_labels = [1] * len(OOD_val_sentences)
    
    # sample_size = min(len(OOD_train_sentences) // 2, 1300)
    # OOD_val_sentences = random.sample(OOD_train_sentences, sample_size)
    # OOD_val_labels = [2] * len(OOD_val_sentences)
    # for item in OOD_val_sentences:
    #     OOD_train_sentences.remove(item)
    

    print('\nOOD_train_sentences', len(OOD_train_sentences))
    print('OOD_test_sentences', len(OOD_test_sentences))
    print('OOD_test_labels', len(OOD_test_labels))
    print('OOD_val_sentences', len(OOD_val_sentences))
    print('OOD_val_labels', len(OOD_val_labels))


    ###Merge IID+generated-OOD dataset
    train_sentences.extend(OOD_train_sentences)
    train_labels.extend(OOD_train_labels)
    test_sentences.extend(OOD_test_sentences)
    test_labels.extend(OOD_test_labels)
    # test_sentences = OOD_test_sentences
    # test_labels= OOD_test_labels
    val_sentences.extend(OOD_val_sentences)
    val_labels.extend(OOD_val_labels)
    
    print('\n\ntrain_labels_IID+OOD', len(train_labels))
    print('Unique train labels', {num: train_labels.count(num) for num in train_labels})
    print('Unique test labels', {num: test_labels.count(num) for num in test_labels})
    print('Unique val labels', {num: val_labels.count(num) for num in val_labels})
    print('test_labels', len(test_labels))
    print('val_labels', len(val_labels))

    return train_sentences, train_labels, test_sentences, test_labels, val_sentences, val_labels


def load_dataset_custom(params):
    """
    Load train and test data
    :param params: experiment parameter, which contains dataset spec
    :return: train_x, train_y, test_x, test_y
    """
    if params['dataset'] == 'sst2':
        orig_train_sentences, orig_train_labels, orig_test_sentences, orig_test_labels, orig_val_sentences, orig_val_labels = load_sst2()
        params['prompt_prefix'] = ""
        params["q_prefix"] = "Review: "
        params["a_prefix"] = "Sentiment: "
        params['label_dict'] = {0: ['Negative'], 1: ['Positive']}
        params['inv_label_dict'] = {'Negative': 0, 'Positive': 1}
        params['task_format'] = 'classification'
        params['num_tokens_to_predict'] = 1

    elif params['dataset'] == 'civil_comments_toxicity_OOD_gsm8k':
        orig_train_sentences, orig_train_labels, orig_test_sentences, orig_test_labels, orig_val_sentences, orig_val_labels = load_civil_comments_toxicity_OOD_gsm8k()
        params['prompt_prefix'] = ""
        params["q_prefix"] = "Review: "
        params["a_prefix"] = "Sentiment: "
        params['label_dict'] = {0: ['Positive'], 1: ['Negative']}
        params['inv_label_dict'] = {'Positive': 0, 'Negative': 1}
        params['task_format'] = 'classification'
        params['num_tokens_to_predict'] = 1

    elif params['dataset'] == 'civil_comments_toxicity_OOD_mbpp':
        orig_train_sentences, orig_train_labels, orig_test_sentences, orig_test_labels, orig_val_sentences, orig_val_labels = load_civil_comments_toxicity_OOD_mbpp()
        params['prompt_prefix'] = ""
        params["q_prefix"] = "Review: "
        params["a_prefix"] = "Sentiment: "
        params['label_dict'] = {0: ['Positive'], 1: ['Negative']}
        params['inv_label_dict'] = {'Positive': 0, 'Negative': 1}
        params['task_format'] = 'classification'
        params['num_tokens_to_predict'] = 1

    elif params['dataset'] == 'civil_comments_toxicity_OOD_sst2':
        orig_train_sentences, orig_train_labels, orig_test_sentences, orig_test_labels, orig_val_sentences, orig_val_labels = load_civil_comments_toxicity_OOD_sst2()
        params['prompt_prefix'] = ""
        params["q_prefix"] = "Review: "
        params["a_prefix"] = "Sentiment: "
        params['label_dict'] = {0: ['Positive'], 1: ['Negative']}
        params['inv_label_dict'] = {'Positive': 0, 'Negative': 1}
        params['task_format'] = 'classification'
        params['num_tokens_to_predict'] = 1

    elif params['dataset'] == 'civil_comments_toxicity_OOD_toxigen':
        orig_train_sentences, orig_train_labels, orig_test_sentences, orig_test_labels, orig_val_sentences, orig_val_labels = load_civil_comments_toxicity_OOD_toxigen()
        params['prompt_prefix'] = ""
        params["q_prefix"] = "Review: "
        params["a_prefix"] = "Sentiment: "
        params['label_dict'] = {0: ['Positive'], 1: ['Negative']}
        params['inv_label_dict'] = {'Positive': 0, 'Negative': 1}
        params['task_format'] = 'classification'
        params['num_tokens_to_predict'] = 1

    elif params['dataset'] == 'response_beavertails_unethical_OOD_gsm8k':
        orig_train_sentences, orig_train_labels, orig_test_sentences, orig_test_labels, orig_val_sentences, orig_val_labels = load_beavertails_unethical_OOD_gsm8k()
        params['prompt_prefix'] = ""
        params["q_prefix"] = " "
        params["a_prefix"] = "Sentiment: "
        params['label_dict'] = {0: ['Positive'], 1: ['Negative']}
        params['inv_label_dict'] = {'Positive': 0, 'Negative': 1}
        params['task_format'] = 'classification'
        params['num_tokens_to_predict'] = 1

    elif params['dataset'] == 'response_beavertails_unethical_OOD_mbpp':
        orig_train_sentences, orig_train_labels, orig_test_sentences, orig_test_labels, orig_val_sentences, orig_val_labels = load_beavertails_unethical_OOD_mbpp()
        params['prompt_prefix'] = ""
        params["q_prefix"] = " "
        params["a_prefix"] = "Sentiment: "
        params['label_dict'] = {0: ['Positive'], 1: ['Negative']}
        params['inv_label_dict'] = {'Positive': 0, 'Negative': 1}
        params['task_format'] = 'classification'
        params['num_tokens_to_predict'] = 1

    elif params['dataset'] == 'response_beavertails_unethical_OOD_sexual-drug':
        orig_train_sentences, orig_train_labels, orig_test_sentences, orig_test_labels, orig_val_sentences, orig_val_labels = load_beavertails_unethical_OOD_sexual_drug()
        params['prompt_prefix'] = ""
        params["q_prefix"] = " "
        params["a_prefix"] = "Sentiment: "
        params['label_dict'] = {0: ['Positive'], 1: ['Negative']}
        params['inv_label_dict'] = {'Positive': 0, 'Negative': 1}
        params['task_format'] = 'classification'
        params['num_tokens_to_predict'] = 1


    elif params['dataset'] == 'response_beavertails_unethical_OOD_discrimincation-hate':
        orig_train_sentences, orig_train_labels, orig_test_sentences, orig_test_labels, orig_val_sentences, orig_val_labels = load_beavertails_unethical_OOD_discrimincation_hate()
        params['prompt_prefix'] = ""
        params["q_prefix"] = " "
        params["a_prefix"] = "Sentiment: "
        params['label_dict'] = {0: ['Positive'], 1: ['Negative']}
        params['inv_label_dict'] = {'Positive': 0, 'Negative': 1}
        params['task_format'] = 'classification'
        params['num_tokens_to_predict'] = 1

    else:
        raise NotImplementedError
    return orig_train_sentences, orig_train_labels, orig_test_sentences, orig_test_labels, orig_val_sentences, orig_val_labels